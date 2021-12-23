# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 10:44:04 2021

"""
import openpyxl
# 引入字体包
from openpyxl.styles import Font
# 设置对齐方式
from openpyxl.styles import Alignment
# 设置线条的形状与颜色
from openpyxl.styles import Side
# 设置边框样式
from openpyxl.styles import Border
# 引入填充单一颜色库
from openpyxl.styles import PatternFill

workbook = openpyxl.load_workbook("./Test表.xlsx")
sheet = workbook["Sheet1"]
# 1.设置字体样式
font = Font(name="微软雅黑", bold=True,italic=True,size="30", color="1FBB7D")
# 2.设置对齐方式
align = Alignment(horizontal="center", vertical="center")
# 3.设置线条的形状与颜色
side = Side(style="thin", color="FF0000")
# 3.设置边框方位样式
border = Border(top=side,bottom=side,left=side,right=side)
# 4.设置单一颜色样式
pattern_fill = PatternFill(fill_type='solid', fgColor='ffff00')

# 将公式应用到单元格中
sheet["F2"] = "=SUM(B2:E2)"
# 将字体样式应用到单元格中
sheet["F2"].font = font
# 将对齐方式应用到单元格中
sheet["F2"].alignment = align
# 将边框样式应用到单元格中
sheet["F2"].border = border
# 将单一填充颜色样式应用到单元格中
sheet["F2"].fill = pattern_fill

# 设置宽度与高
sheet.row_dimensions[2].height = 100
sheet.column_dimensions['F'].width = 100

# 合并单元格
sheet.merge_cells("B2:E2")
# 取消单元格
sheet.unmerge_cells("B4:E4")
# 冻结窗格
sheet.freeze_panes = 'A2'
# 添加帅选
sheet.auto_filter.ref = sheet.dimensions

workbook.save("./Test表哥.xlsx")








